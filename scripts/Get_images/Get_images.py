# Importing the modules
import openpyxl
from openpyxl_image_loader import SheetImageLoader
import os
import unidecode


# Loading the Excel File and the sheet for filenames
pxl_doc_filenames = openpyxl.load_workbook('GES267-Studies.xlsx')
sheet_filenames = pxl_doc_filenames['Gesture Elicitation Studies']

# Loading the Excel File and the sheet for ID check
pxl_doc_id = openpyxl.load_workbook('GES267-GesturesV2.xlsx')
sheet_id = pxl_doc_id['RefGes']


# Loading the Excel File and the sheet for images
pxl_doc_images = openpyxl.load_workbook('GES267-GesturesV2.xlsx')
sheet_images = pxl_doc_images['RefGes']

# Calling the image_loader
image_loader = SheetImageLoader(sheet_images)


# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or column integer is 1, not 0.

def get_author_name(j):
    # Author_name
    # Cell object is created by using sheet object's cell() method.
    cell_obj = sheet_filenames.cell(row=j, column=4)

    # Get value of cell object using the value attribute
    author_name_liste = cell_obj.value.replace(",", " ").replace(";", " ").replace(".", " ").split()
    author_name = ""
    for i in range(len(author_name_liste)):
        if (len(author_name_liste[i]) > 2):
            author_name = author_name_liste[i].lower()
            author_name = unidecode.unidecode(author_name)
            break
    return author_name

def get_title(j):
    # Title
    cell_obj = sheet_filenames.cell(row=j, column=6)

    # Get value of cell object using the value attribute
    title_list = cell_obj.value.split()
    title = title_list[0].lower()
    i = 1
    while (len(title) <= 3):
        title = title_list[i].lower()
        i += 1
    title = title.replace(":", "")
    return title


def get_year(j):
    # Year
    cell_obj = sheet_filenames.cell(row=j, column=3)

    # Get value of cell object using the value attribute
    year = cell_obj.value

    return year


def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print('Error: Creating directory. ' + directory)


def get_images(cell, id_flag, flag_check, number):
    image = image_loader.get(cell)
    filename = get_author_name(id_flag) + "_" + get_title(id_flag) + "_" + str(get_year(id_flag))
    if (flag_check):
        createFolder(filename)
    image_name = filename + "/" + number + ".png"
    image_location = "data/imgGesture/" + image_name
    with open("Images_location.txt", "a") as file:
        file.write(image_location + "\n")
    # saving the image
    image.save(image_name)

def image_check(cell):
    not_image = False
    try:
        image = image_loader.get(cell)
    except ValueError:
        not_image = True
    if (not_image):
        return False
    else:
        return True


id_flag = 0
count = 1
directory_created = False
# Get all images in the right directory
for i in range (2, 3912):
#for i in range(2, 200):
    cell_obj_id = sheet_id.cell(row=i, column=1).value
    cell = 'F' + str(i)
    if (cell_obj_id != None and cell_obj_id > id_flag):
        count = 1
        str_count = "0" + str(count)
        id_flag += 1
        if (image_check(cell)):
            get_images(cell, id_flag + 1, True, str_count)
            directory_created = True
        else:
            directory_created = False
    else:
        if (image_check(cell)):
            if (directory_created):
                count += 1
            if (count < 10):
                str_count = "0" + str(count)
            else:
                str_count = str(count)
            if (directory_created):
                get_images(cell, id_flag + 1, False, str_count)
            else:
                get_images(cell, id_flag + 1, True, str_count)
                directory_created = True
        else:
            continue